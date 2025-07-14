# views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from datetime import timedelta
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.contrib import messages
from django.contrib.auth.models import User, Group
from .models import Karyawan, Benefit, Kedisiplinan, SuratPeringatan, PenerimaBenefit
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.db.models import Count
from django.utils.timezone import now
from .models import Benefit
import pandas as pd
from django.contrib import messages
from .models import Karyawan


# ⬅️ Pindahkan ke atas
def is_admin(user):
    return user.groups.filter(name='Administrator').exists() or user.is_superuser

# Toggle status pengguna
@user_passes_test(is_admin)
def toggle_status_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()
    messages.success(request, f"Status pengguna '{user.username}' berhasil diperbarui.")
    return redirect('kelola_pengguna')


# Halaman Kelola Pengguna################################################################################################################
@user_passes_test(is_admin)
def kelola_pengguna(request):
    users = User.objects.all().order_by('username')
    groups = Group.objects.all()
    return render(request, 'akses/pengguna_list.html', {'users': users, 'groups': groups})


@user_passes_test(is_admin)
def update_grup_pengguna(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        group_name = request.POST.get('group')
        if group_name:
            user.groups.clear()
            user.groups.add(Group.objects.get(name=group_name))
        return redirect('kelola_pengguna')
    return redirect('kelola_pengguna')


#dashboard--###############################################################################################################
@login_required
def dashboard(request):
    today = timezone.now().date()
    total_benefit_today = Benefit.objects.filter(tanggal=today).count()
    total_karyawan = Karyawan.objects.count()
    total_sp_aktif = SuratPeringatan.objects.filter(tanggal_berlaku_sampai__gte=today).count()
    total_karyawan_bawa_hp = Karyawan.objects.filter(akses_hp=True).count()
    total_user = User.objects.filter(is_active=True).count()

    return render(request, 'dashboard.html', {
        'total_benefit_today': total_benefit_today,
        'total_karyawan': total_karyawan,
        'total_sp_aktif': total_sp_aktif,
        'total_izin_bawa_hp': total_karyawan_bawa_hp,
        'total_user': total_user
    })

#informasi--###############################################################################################################
# Akses berdasarkan permission
@permission_required('core.can_scan_informasi')
def informasi_page(request):
    return render(request, 'scan/informasi.html')

@user_passes_test(is_admin)
def karyawan_list(request):
    karyawan = Karyawan.objects.all()
    return render(request, 'rekap/karyawan_list.html', {'karyawan': karyawan})

def api_informasi(request, no_id):
    try:
        karyawan = Karyawan.objects.get(no_id=no_id)
        data = {
            "nama": karyawan.nama,
            "departemen": karyawan.departemen,
            "nik": karyawan.nik,
            "bpjs_k": karyawan.bpjs_k,
            "bpjs_tk": karyawan.bpjs_tk,
            "no_hp": karyawan.no_hp,
            "akses_hp": "Ya" if karyawan.akses_hp else "Tidak",
            "akses_laptop": "Ya" if karyawan.akses_laptop else "Tidak",
        }
        return JsonResponse({"status": "ok", "data": data})
    except Karyawan.DoesNotExist:
        return JsonResponse({"status": "notfound"})

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrator').exists())
def karyawan_tambah(request):
    if request.method == 'POST':
        no_id = request.POST.get('no_id')
        nama = request.POST.get('nama')
        departemen = request.POST.get('departemen')
        nik = request.POST.get('nik')
        bpjs_k = request.POST.get('bpjs_k')
        bpjs_tk = request.POST.get('bpjs_tk')
        no_hp = request.POST.get('no_hp')
        akses_hp = bool(request.POST.get('akses_hp'))
        akses_laptop = bool(request.POST.get('akses_laptop'))

        Karyawan.objects.create(
            no_id=no_id,
            nama=nama,
            departemen=departemen,
            nik=nik,
            bpjs_k=bpjs_k,
            bpjs_tk=bpjs_tk,
            no_hp=no_hp,
            akses_hp=akses_hp,
            akses_laptop=akses_laptop
        )
        messages.success(request, "Data karyawan berhasil ditambahkan.")
        return redirect('karyawan_list')
    
    return render(request, 'rekap/karyawan_tambah.html')

@login_required
@user_passes_test(is_admin)
def upload_karyawan_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        df = pd.read_excel(request.FILES['excel_file'])

        for _, row in df.iterrows():
            Karyawan.objects.update_or_create(
                no_id=row.get('no_id'),
                defaults={
                    'nama': row.get('nama'),
                    'departemen': row.get('departemen'),
                    'nik': row.get('nik'),
                    'bpjs_k': row.get('bpjs_k'),
                    'bpjs_tk': row.get('bpjs_tk'),
                    'no_hp': row.get('no_hp'),
                    'akses_hp': bool(row.get('akses_hp')),
                    'akses_laptop': bool(row.get('akses_laptop')),
                }
            )
        messages.success(request, "Data karyawan berhasil diunggah.")
        return redirect('karyawan_list')

    return render(request, 'rekap/karyawan_upload.html')


#benefit-###############################################################################################################

@permission_required('core.can_scan_benefit')
def benefit_page(request):
    return render(request, 'scan/benefit.html')


def api_benefit(request, no_id):
    try:
        karyawan = Karyawan.objects.get(no_id=no_id)
        penerima = PenerimaBenefit.objects.filter(karyawan=karyawan)

        if not penerima.exists():
            return JsonResponse({'status': 'not_eligible'})

        jenis_list = list(penerima.values_list('jenis', flat=True))

        riwayat = Benefit.objects.filter(karyawan=karyawan).order_by('-tanggal')[:5]
        riwayat_data = list(riwayat.values('jenis', 'tanggal'))

        return JsonResponse({
            'status': 'eligible',
            'data': {
                'nama': karyawan.nama,
                'departemen': karyawan.departemen,
                'no_id': karyawan.no_id,
                'jenis': jenis_list,
                'riwayat': riwayat_data
            }
        })
    except Karyawan.DoesNotExist:
        return JsonResponse({'status': 'not_found'})

@csrf_exempt
def take_benefit(request):
    if request.method == 'POST':
        no_id = request.POST.get('no_id')
        jenis = request.POST.get('jenis')

        try:
            karyawan = Karyawan.objects.get(no_id=no_id)

            if not PenerimaBenefit.objects.filter(karyawan=karyawan, jenis=jenis).exists():
                return JsonResponse({'success': False, 'message': 'Tidak berhak menerima benefit'})

            if Benefit.objects.filter(karyawan=karyawan, jenis=jenis).exists():
                return JsonResponse({'success': False, 'message': 'Sudah pernah ambil'})

            Benefit.objects.create(karyawan=karyawan, jenis=jenis)
            return JsonResponse({'success': True})

        except Karyawan.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'ID tidak ditemukan'})

        

@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name="Administrator").exists())
def rekap_benefit(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data = Benefit.objects.select_related('karyawan').order_by('-tanggal')

    if start_date and end_date:
        data = data.filter(tanggal__range=[start_date, end_date])

    return render(request, 'rekap/benefit.html', {
        'data': data,
        'start_date': start_date,
        'end_date': end_date
    })

@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name="Administrator").exists())
def export_benefit_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data = Benefit.objects.select_related('karyawan').order_by('-tanggal')

    if start_date and end_date:
        data = data.filter(tanggal__range=[start_date, end_date])

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Benefit"

    headers = ['Tanggal', 'No ID', 'Nama', 'Departemen', 'Benefit']
    ws.append(headers)

    for b in data:
        ws.append([
            b.tanggal.strftime('%Y-%m-%d'),
            b.karyawan.no_id,
            b.karyawan.nama,
            b.karyawan.departemen,
            b.jenis
        ])

    # Autofit column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Output response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"rekap_benefit_{start_date or 'awal'}_to_{end_date or 'akhir'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

def benefit_list(request):
    data_taken = Benefit.objects.select_related('karyawan').order_by('-tanggal')
    data_penerima = PenerimaBenefit.objects.select_related('karyawan')

    return render(request, 'rekap/benefit.html', {
        'data_taken': data_taken,
        'data_penerima': data_penerima
    })


@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrator').exists())
def upload_benefit_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        df = pd.read_excel(request.FILES['excel_file'])

        for _, row in df.iterrows():
            try:
                karyawan = Karyawan.objects.get(no_id=row.get('no_id'))
                Benefit.objects.create(
                    karyawan=karyawan,
                    jenis=row.get('jenis'),
                    keterangan=row.get('keterangan'),
                    tanggal=None  # default kosong
                )
            except Karyawan.DoesNotExist:
                continue

        messages.success(request, "Data Benefit berhasil diunggah.")
        return redirect('benefit_list')  # sesuaikan jika URL list benefit berbeda

    return redirect('benefit_list')

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrator').exists())
def upload_penerima_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        df = pd.read_excel(request.FILES['excel_file'])
        print("Header kolom Excel:", df.columns)  # debug

        for _, row in df.iterrows():
            try:
                print("Proses:", row.get('no_id'), row.get('jenis'))
                karyawan = Karyawan.objects.get(no_id=row.get('no_id'))
                PenerimaBenefit.objects.update_or_create(
                    karyawan=karyawan,
                    jenis=row.get('jenis'),
                    defaults={'keterangan': row.get('keterangan')}
                )
            except Karyawan.DoesNotExist:
                print("Karyawan tidak ditemukan:", row.get('no_id'))
                continue
            except Exception as e:
                print("Error lain:", e)

        messages.success(request, "Data penerima benefit berhasil diunggah.")
        return redirect('benefit_list')

    return redirect('benefit_list')
#-----------------------------------------------------------------------------------------------------------
@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrator').exists())
def upload_sp_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        df = pd.read_excel(request.FILES['excel_file'])

        for _, row in df.iterrows():
            try:
                karyawan = Karyawan.objects.get(no_id=row.get('no_id'))
                SuratPeringatan.objects.create(
                    karyawan=karyawan,
                    jenis=row.get('jenis_sp'),
                    tanggal_terbit=row.get('tanggal_terbit'),
                    tanggal_berlaku_sampai=row.get('tanggal_berlaku_sampai'),
                    keterangan=row.get('keterangan')
                )
            except Karyawan.DoesNotExist:
                continue

        messages.success(request, "Data Surat Peringatan berhasil diunggah.")
        return redirect('sp_list')
    
    return render(request, 'rekap/sp_upload.html')

#kedisiplinan-------------------------------------------------------------------------------------------------
@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrator').exists())
def upload_kedisiplinan_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        df = pd.read_excel(request.FILES['excel_file'])

        for _, row in df.iterrows():
            try:
                karyawan = Karyawan.objects.get(no_id=row.get('no_id'))
                Kedisiplinan.objects.create(
                    karyawan=karyawan,
                    tanggal=row.get('tanggal'),
                    jenis=row.get('jenis_kedisiplinan'),
                    keterangan=None  # atau bisa ambil dari kolom jika ada
                )
            except Karyawan.DoesNotExist:
                continue

        messages.success(request, "Data Kedisiplinan berhasil diunggah.")
        return redirect('sp_list')  # atau halaman lain sesuai kebutuhan
    
    return render(request, 'rekap/kedisiplinan_upload.html')


@permission_required('core.can_scan_kedisiplinan')
def kedisiplinan_page(request):
    return render(request, 'scan/kedisiplinan.html')

def api_kedisiplinan(request, no_id):
    try:
        karyawan = Karyawan.objects.get(no_id=no_id)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=90)

        records = Kedisiplinan.objects.filter(karyawan=karyawan, tanggal__range=(start_date, end_date))
        sp_aktif = SuratPeringatan.objects.filter(
            karyawan=karyawan,
            tanggal_berlaku_sampai__gte=end_date
        ).order_by('-tanggal_terbit')

        return JsonResponse({
            'status': 'ok',
            'data': {
                'nama': karyawan.nama,
                'departemen': karyawan.departemen,
                'A': records.filter(jenis='A').count(),
                'DT': records.filter(jenis='DT').count(),
                'PC': records.filter(jenis='PC').count(),
                'CX': records.filter(jenis='CX').count(),
                'sp': list(sp_aktif.values('jenis', 'tanggal_terbit', 'tanggal_berlaku_sampai')),
            }
        })

    except Karyawan.DoesNotExist:
        return JsonResponse({'status': 'not_found'})
    
@login_required
def sp_list(request):
    from django.utils import timezone
    today = timezone.now().date()
    start_date = today - timedelta(days=90)

    sp = SuratPeringatan.objects.all().order_by('-tanggal_terbit')
    kedisiplinan = Kedisiplinan.objects.filter(tanggal__range=(start_date, today)).order_by('-tanggal')

    return render(request, 'rekap/sp_list.html', {
        'sp': sp,
        'kedisiplinan': kedisiplinan
    })
